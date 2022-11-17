from univer.models import DirectionTraining, StudyGroup

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from celery import shared_task


@shared_task
def report_exel():
    report = Workbook()
    report.remove(report.active)
    """Формирование первого листа отчета"""
    sheet_direction_training = report.create_sheet(title='Информация о направлениях', index=1)
    lst_direction_training = DirectionTraining.objects.all()
    columns = ['Title', 'Content', 'Academic Discipline', 'Tutor', 'Age', 'Phone']
    row_num = 1
    """Заполнение заголовков первого листа"""
    for col_num, column_title in enumerate(columns, 1):
        cell = sheet_direction_training.cell(row=row_num, column=col_num)
        cell.value = column_title
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.font = Font(bold=True)

    """Формирование и заполнение строк первого листа"""
    for i, direction_training in enumerate(lst_direction_training, 1):
        row_num += 1
        row = [
            direction_training.title,
            direction_training.content,
            ', '.join(list(i.title for i in direction_training.academ_dis.all())),
            ', '.join(list(i.username for i in direction_training.tutors.filter(is_tutor=True))),
            ', '.join(list(str(i.age) for i in direction_training.tutors.filter(is_tutor=True))),
            ', '.join(list(i.phone for i in direction_training.tutors.filter(is_tutor=True))),
        ]
        for col_num, cell_value in enumerate(row, 1):
            cell = sheet_direction_training.cell(row=row_num, column=col_num)
            cell.value = cell_value

    """Формирование второго листа отчета"""
    sheet_study_group = report.create_sheet(title='Информация о группах', index=2)
    lst_study_group = StudyGroup.objects.all()
    columns_study_group = ['Title', 'Students', 'Count men', 'Count women', 'Free seats']

    """Заполнение заголовков второго листа"""
    for col_num, column_title in enumerate(columns_study_group, 1):
        cell = sheet_study_group.cell(row=row_num, column=col_num)
        cell.value = column_title
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.font = Font(bold=True)

    """Формирование и заполнение строк второго листа"""
    for i, study_group in enumerate(lst_study_group, 1):
        row_num += 1
        row = [
            study_group.title,
            ', '.join(sorted(list(i.username for i in study_group.students.all()))),
            study_group.students.filter(gender_id=1).count(),
            study_group.students.filter(gender_id=2).count(),
            3 - int(study_group.students.all().count()),
        ]
        for col_num, cell_value in enumerate(row, 1):
            cell = sheet_study_group.cell(row=row_num, column=col_num)
            cell.value = cell_value

    report.save('report.xlsx')
    report.close()
